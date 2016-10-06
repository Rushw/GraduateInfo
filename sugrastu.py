# -*- coding: utf-8 -*-
import urllib,urllib2, multiprocessing, datetime
import sys
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
from zipfile import ZipFile, ZIP_DEFLATED


class InfoExporter(object):
    extension = ".xlsx"
    from openpyxl.styles import colors, Font
    hyperlink_font = Font(underline="single", color=colors.BLUE)
    bold_font = Font(bold=True)
    
    def __init__(self, year):
        self._workbook = Workbook()
        self.year = year
        from openpyxl.writer.excel import ExcelWriter
        self._excel_writer = ExcelWriter(workbook=self._workbook,archive=ZipFile('%s.zip'%self.year,"a",ZIP_DEFLATED))
        self.add_sheet(self.year)
        
        
        
    def add_sheet(self, title):
        sheet = self._workbook.create_sheet(title=title)
        self.titles = [
            {"name": "学号", "width": 10,'id':'lblxh'},
            {"name": "校内编号", "width": 10,'id':'lblxnbh'},
            {"name": "姓名", "width": 10,'id':'lblxm'},
            {"name": "性别", "width": 5,'id':'lblxb'},
            {"name": "籍贯", "width": 40,'id':'lbljg'},
            {"name": "出生年份", "width": 10,'id':'lblcsrq'},
            {"name": "民族", "width": 10,'id':'lblmz'},
            {"name": "婚姻状况", "width": 10,'id':'lblhyzk'},
            {"name": "政治面貌", "width": 10,'id':'lblzzmm'},
            {"name": "入党年月", "width": 20,'id':'lblrdny'},
            {"name": "入学年月", "width": 20,'id':'lblrxny'},
            {"name": "考生来源", "width": 10,'id':'lblksly'},
            {"name": "家庭地区码", "width": 15,'id':'lbljtdqm'},
            {"name": "家庭地区", "width": 30,'id':'lbljtdq'},
            {"name": "入学方式", "width": 10,'id':'lblrxfs'},
            {"name": "入学前最后学历毕业院校码", "width": 30,'id':'lblbyyxm'},
            {"name": "最后毕业院校", "width": 30,'id':'lblbyyx'},
            {"name": "入学前最后学历毕业年月", "width": 30,'id':'lblybyrq'},
            {"name": "入学前工作单位", "width": 30,'id':'lblygzdw'},
            {"name": "是否留学人员", "width": 20,'id':'lblsflx'},
            {"name": "培养类别", "width": 10,'id':'lblpylb'},
            {"name": "专业", "width": 20,'id':'lblzymc'},
            {"name": "导师", "width": 10,'id':'lbldsxm'},
            {"name": "学位类型", "width": 20,'id':'lblxwlx'},
            {"name": "研究方向", "width": 30,'id':'lblyjfx'},
            {"name": "论文题目", "width": 100,'id':'lbllwtm'},
            {"name": "论文起止日期", "width": 60,'id':'lbllwqzrq'},
            {"name": "答辩日期", "width": 30,'id':'lbldbrq'},
            {"name": "学位证号", "width": 20,'id':'lblxwzh'},
            {"name": "授学位日期", "width": 60,'id':'lblsxwrq'},
            {"name": "院系", "width": 30,'id':'lblyx'},
            {"name": "毕业证号", "width": 20,'id':'lblbyzh'},
            {"name": "毕业时间", "width": 20,'id':'lblbysj'},
            {"name": "分配单位", "width": 20, 'id': 'lblfpdw'},
            {"name": "分配单位类别", "width": 20,'id':'lblfpdwlb'},
            {"name": "学籍异动标记", "width": 20,'id':'lblxjyd'},
            {"name": "奖惩标记", "width": 10,'id':'lbljcbj'},
            {"name": "结束学业码", "width": 15,'id':'lbljsxy'},
            {"name": "结束学业年月", "width": 30,'id':'lbljsxyny'},
            {"name": "备注", "width": 20,'id':'lblbz'},
        ]
        for col in range(0, len(self.titles)):
            cell = sheet.cell('%s%s' % (get_column_letter(col + 1), 1))
            cell.value = self.titles[col]["name"]
            cell.font = self.bold_font
            sheet.column_dimensions[
                get_column_letter(col + 1)].width = int(self.titles[col]["width"])
        self._workbook.active = self._workbook.get_index(sheet)
    
    def _get_cell_string(self, col, row):
        return '%s%s' % (get_column_letter(col), row)
    
    def report(self):
        sheet = self._workbook.active
        initialRow = rowNumber = sheet.max_row + 1
        nullcounter = 0    # avoid continuing query the null ID section
        failnums = []
        for num in range(1, 0005):
            print 'processing information for student ID %s%s'% (self.year, str(num).zfill(4))
            try:
                responce = urllib.urlopen(
                    'http://202.119.4.150/nstudent/ggxx/xsggxxinfo.aspx?xh=%s%s' % (self.year, str(num).zfill(4)),
                )
            except:
                failnums.append(num)
                continue
            text = responce.read().decode('utf-8')
            soup = BeautifulSoup(text,'html.parser')
            if str(soup.find('span', {'id': 'lblxh'}).text).isdigit():
                nullcounter=0
                for i,propty in enumerate(self.titles):
                    sheet.cell(self._get_cell_string(i+1, rowNumber)
                               ).value = soup.find('span',{'id':propty['id']}).text
                rowNumber+=1
            else:
                nullcounter+=1
                if nullcounter>20:
                    break
        for num in failnums: # retry for timeout issue.
            print 'processing information for student ID %s%s' % (self.year, str(num).zfill(4))
            try:
                responce = urllib.urlopen(
                    'http://202.119.4.150/nstudent/ggxx/xsggxxinfo.aspx?xh=%s%s' % (self.year, str(num).zfill(4)),
                )
            except:
                failnums.append(num)
                continue
            text = responce.read().decode('utf-8')
            soup = BeautifulSoup(text, 'html.parser')
            if str(soup.find('span', {'id': 'lblxh'}).text).isdigit():
                nullcounter = 0
                for i, propty in enumerate(self.titles):
                    sheet.cell(self._get_cell_string(i + 1, rowNumber)
                               ).value = soup.find('span', {'id': propty['id']}).text
                rowNumber += 1
            else:
                nullcounter += 1
                if nullcounter > 20:
                    break
    def save(self):
        from openpyxl.writer.excel import save_workbook
        save_workbook(self._workbook,self.year + ".xlsx")


def get_all((year)):
    reporter = InfoExporter(year)
    reporter.report()
    reporter.save()

def main():
    args = sys.argv[1:]
    if len(args) != 2:
        print 'usage: start end'
        print 'start and end should in xx format. pls ignore 20 in 20xx'
        sys.exit(1)
    if int(args[0]) < int(args[1]) and int(args[0]) > 0 and (2000 + int(args[1])) > datetime.datetime.now().year:
        print 'constraint: start >= 00, end <= year-2000.'
        sys.exit(1)
    procs = multiprocessing.Pool()
    pargs = []
    for i in range(int(args[0]), int(args[1])):
        year = str(i).zfill(2)
        # get_all(year)
        pargs.append(year)
    procs.map(get_all,pargs)
    procs.close()
    procs.join()
    print 'all information is logged into xx.xlsx'


if __name__ == '__main__':
    main()
